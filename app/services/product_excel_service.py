from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.money import display, display_qty, money, qty
from app.db.database import Database
from app.services.products_service import ProductsService


PRODUCT_EXCEL_HEADERS: list[str] = [
    "اسم المنتج *",
    "الباركود",
    "رمز مسح إضافي",
    "الفئة",
    "المورد",
    "وحدة القياس",
    "هل المنتج موزون؟ (1/0)",
    "هل له أجزاء؟ (1/0)",
    "اسم الجزء",
    "عدد الأجزاء في الوحدة",
    "سعر الشراء USD *",
    "سعر البيع مفرق USD *",
    "سعر البيع جملة USD",
    "حد الجملة (كمية)",
    "الكمية الافتتاحية",
    "الحد الأدنى للمخزون",
    "تاريخ الصلاحية (YYYY-MM-DD أو DD/MM/YYYY أو MM/DD/YYYY)",
    "ملاحظات",
    "رابط الصورة (image_url)",
]

EXAMPLE_PRODUCT_ROW: list[Any] = [
    "مثال: رز بسمتي",
    "مثال: 6281234567890",
    "مثال: QR-6281234567890",
    "اعشاب | الكراسي | خزن | طاولات | غرف النوم | فرش",
    "سمير",
    "قطعة | كغ | غ | لتر | مل | علبة | كرتون | حزمة | متر | دزينة",
    "0",
    "1",
    "مثال: ظرف",
    "12",
    "5.50",
    "8.00",
    "7.00",
    "10",
    "100",
    "5",
    "2025-12-31",
    "أي ملاحظة إضافية",
    "اختياري: https://example.com/product.jpg أو data:image/png;base64,...",
]


class ProductExcelService:
    def __init__(self, db: Database):
        self.db = db
        self.products = ProductsService(db)

    def import_excel(self, path: str | Path) -> dict[str, Any]:
        file_path = Path(path)
        if not file_path.exists():
            raise FileNotFoundError(f"ملف الإكسل غير موجود: {file_path}")

        workbook = load_workbook(file_path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()
        if not rows:
            raise ValueError("ملف الإكسل فارغ")

        header_row = [self._normalize_header(v) for v in rows[0]]
        missing = [h for h in PRODUCT_EXCEL_HEADERS if h not in header_row]
        if missing:
            raise ValueError("أعمدة ناقصة في ملف المنتجات:\n" + "\n".join(missing[:8]))

        indexes = {h: header_row.index(h) for h in PRODUCT_EXCEL_HEADERS}
        created = 0
        updated = 0
        skipped = 0
        errors: list[str] = []

        for excel_row_number, row in enumerate(rows[1:], start=2):
            try:
                data = self._row_to_product_data(row, indexes)
                name = str(data.get("name") or "").strip()
                if not name or name.startswith("مثال:"):
                    skipped += 1
                    continue

                existing_id = self._find_existing_product_id(data)
                if existing_id:
                    self.products.update_product(existing_id, data)
                    updated += 1
                else:
                    self.products.create_product(data)
                    created += 1
            except Exception as exc:
                errors.append(f"السطر {excel_row_number}: {exc}")

        return {"created": created, "updated": updated, "skipped": skipped, "errors": errors}

    def export_excel(self, path: str | Path) -> Path:
        out = Path(path)
        if out.suffix.lower() != ".xlsx":
            out = out.with_suffix(".xlsx")
        out.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "المنتجات"
        sheet.sheet_view.rightToLeft = True

        sheet.append(PRODUCT_EXCEL_HEADERS)
        sheet.append(EXAMPLE_PRODUCT_ROW)

        for product in self.products.list_products("", active_only=False, limit=100000):
            sheet.append(self._product_to_excel_row(product))

        self._style_sheet(sheet)
        workbook.save(out)
        workbook.close()
        return out

    def _find_existing_product_id(self, data: dict[str, Any]) -> int | None:
        barcode = str(data.get("barcode") or "").strip()
        name = str(data.get("name") or "").strip()
        with self.db.connect() as conn:
            if barcode:
                row = conn.execute("SELECT id FROM products WHERE barcode = ? LIMIT 1", (barcode,)).fetchone()
                if row:
                    return int(row["id"])
            if name:
                row = conn.execute("SELECT id FROM products WHERE name = ? LIMIT 1", (name,)).fetchone()
                if row:
                    return int(row["id"])
        return None

    def _row_to_product_data(self, row: tuple[Any, ...], indexes: dict[str, int]) -> dict[str, Any]:
        def get(header: str) -> Any:
            index = indexes[header]
            return row[index] if index < len(row) else None

        return {
            "name": self._text(get("اسم المنتج *")),
            "barcode": self._text(get("الباركود")),
            "extra_scan_code": self._text(get("رمز مسح إضافي")),
            "category": self._text(get("الفئة")),
            "supplier": self._text(get("المورد")),
            "unit": self._text(get("وحدة القياس")) or "قطعة",
            "is_weighted": self._flag(get("هل المنتج موزون؟ (1/0)")),
            "has_parts": self._flag(get("هل له أجزاء؟ (1/0)")),
            "part_name": self._text(get("اسم الجزء")),
            "parts_per_unit": self._number_text(get("عدد الأجزاء في الوحدة"), default="1"),
            "purchase_price": self._number_text(get("سعر الشراء USD *"), default="0"),
            "retail_price": self._number_text(get("سعر البيع مفرق USD *"), default="0"),
            "wholesale_price": self._number_text(get("سعر البيع جملة USD"), default="0"),
            "wholesale_min_qty": self._number_text(get("حد الجملة (كمية)"), default="1"),
            "stock_quantity": self._number_text(get("الكمية الافتتاحية"), default="0"),
            "min_stock_level": self._number_text(get("الحد الأدنى للمخزون"), default="0"),
            "expiry_date": self._date_text(get("تاريخ الصلاحية (YYYY-MM-DD أو DD/MM/YYYY أو MM/DD/YYYY)")),
            "notes": self._text(get("ملاحظات")),
            "image_url": self._text(get("رابط الصورة (image_url)")),
        }

    def _product_to_excel_row(self, product: dict[str, Any]) -> list[Any]:
        return [
            product.get("name") or "",
            product.get("barcode") or "",
            product.get("extra_scan_code") or "",
            product.get("category") or "",
            product.get("supplier") or "",
            product.get("unit") or "قطعة",
            int(product.get("is_weighted") or 0),
            int(product.get("has_parts") or 0),
            product.get("part_name") or "",
            self._display_number(product.get("parts_per_unit"), quantity=True, fallback="1"),
            self._display_number(product.get("purchase_price"), fallback="0"),
            self._display_number(product.get("retail_price"), fallback="0"),
            self._display_number(product.get("wholesale_price"), fallback="0"),
            self._display_number(product.get("wholesale_min_qty"), quantity=True, fallback="1"),
            self._display_number(product.get("stock_quantity"), quantity=True, fallback="0"),
            self._display_number(product.get("min_stock_level"), quantity=True, fallback="0"),
            product.get("expiry_date") or "",
            product.get("notes") or "",
            product.get("image_url") or "",
        ]

    def _style_sheet(self, sheet) -> None:
        header_fill = PatternFill("solid", fgColor="0F172A")
        example_fill = PatternFill("solid", fgColor="EEF2FF")
        header_font = Font(color="FFFFFF", bold=True)
        example_font = Font(color="475569", italic=True)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cell in sheet[2]:
            cell.fill = example_fill
            cell.font = example_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        widths = [24, 18, 20, 18, 18, 16, 18, 16, 16, 18, 16, 18, 16, 16, 16, 18, 32, 36, 44]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.freeze_panes = "A3"
        sheet.auto_filter.ref = sheet.dimensions

    @staticmethod
    def _normalize_header(value: Any) -> str:
        return str(value or "").replace("\ufeff", "").strip()

    @staticmethod
    def _text(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _flag(value: Any) -> int:
        text = str(value or "").strip().lower()
        return 1 if text in {"1", "true", "yes", "نعم", "صح"} else 0

    @staticmethod
    def _number_text(value: Any, *, default: str) -> str:
        if value is None or str(value).strip() == "":
            return default
        return str(value).strip()

    @staticmethod
    def _date_text(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        return str(value).strip()

    @staticmethod
    def _display_number(value: Any, *, quantity: bool = False, fallback: str) -> str:
        if value is None or str(value).strip() == "":
            return fallback
        try:
            return display_qty(qty(value)) if quantity else display(money(value))
        except Exception:
            return str(value)
