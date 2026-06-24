from __future__ import annotations

import shutil
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.db.database import Database
from app.services.backup_service import BackupService
from app.services.product_excel_service import PRODUCT_EXCEL_HEADERS, ProductExcelService
from app.services.products_service import ProductsService
from app.services.sales_service import SalesService
from app.services.settings_service import SettingsService


def main() -> None:
    temp = Path(tempfile.mkdtemp(prefix="cashier_lite_smoke_"))
    try:
        db = Database(temp / "smoke.sqlite")
        db.initialize()
        products = ProductsService(db)
        sales = SalesService(db)
        settings = SettingsService(db)
        settings.ensure_defaults()
        product_id = products.create_product(
            {
                "name": "منتج تجربة",
                "barcode": "TST-001",
                "unit": "قطعة",
                "category": "تجارب",
                "supplier": "مورد تجربة",
                "purchase_price": "10",
                "retail_price": "15",
                "wholesale_price": "13",
                "stock_quantity": "5",
                "min_stock_level": "1",
            }
        )
        result = sales.create_sale([{"product_id": product_id, "quantity": "2", "unit_price": "15", "discount": "0"}])
        product = products.get_product(product_id)
        assert product is not None
        assert product["stock_quantity"] == "3.0000"
        assert product["category"] == "تجارب"
        assert result["total_amount"] == "30.0000"
        details = sales.get_sale_details(int(result["sale_id"]))
        assert str(details["sale"].get("gross_profit")) in {"10", "10.0", "10.0000"}

        excel = ProductExcelService(db)
        import_file = temp / "products_import.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "المنتجات"
        ws.append(PRODUCT_EXCEL_HEADERS)
        ws.append([
            "منتج مستورد",
            "IMP-001",
            "ALT-IMP-001",
            "فئة مستوردة",
            "مورد مستورد",
            "قطعة",
            0,
            0,
            "",
            1,
            20,
            35,
            30,
            5,
            9,
            2,
            "2027-12-31",
            "ملاحظة استيراد",
            "https://example.com/p.png",
        ])
        wb.save(import_file)
        wb.close()
        import_result = excel.import_excel(import_file)
        assert import_result["created"] == 1
        imported = products.list_products("IMP-001")
        assert len(imported) == 1
        assert imported[0]["category"] == "فئة مستوردة"
        export_file = excel.export_excel(temp / "products_export.xlsx")
        assert export_file.exists()
        wb2 = load_workbook(export_file, read_only=True)
        ws2 = wb2.active
        headers = [cell.value for cell in next(ws2.iter_rows(min_row=1, max_row=1))]
        wb2.close()
        assert headers == PRODUCT_EXCEL_HEADERS

        backup = BackupService(db).create_backup(label="smoke")
        assert backup.exists()
        assert settings.get("theme_mode") == "light"
        print("SMOKE_OK", result["invoice_number"])
    finally:
        shutil.rmtree(temp, ignore_errors=True)


if __name__ == "__main__":
    main()
